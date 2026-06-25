#!/usr/bin/env python3
"""
OpenRouter free-model benchmark runner.

One run = for each enabled category (code / text / email / image), obtain a prompt
(generate it with an AI model, use a manual prompt, or use a fallback seed) and send
that SAME prompt to every configured model. For each model call we capture tokens in,
tokens out, latency, response length and the full response, then write everything to:

  - results.xlsx            full archive, one sheet per category (metrics + preview + file path)
  - results.json            last N runs, consumed by the dashboard
  - responses/<run-id>/...  full response text per model
  - images/<run-id>/...     generated images (only when the image category is enabled)

All behaviour is driven by config.json. See README.md.

Env:
  OPENROUTER_API_KEY   required, your OpenRouter API key
"""

import base64
import json
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import requests
from openpyxl import Workbook, load_workbook

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(ROOT, "config.json")
STATE_PATH = os.path.join(ROOT, "state.json")
HISTORY_PATH = os.path.join(ROOT, "prompt_history.json")
RESULTS_XLSX = os.path.join(ROOT, "results.xlsx")
RESULTS_JSON = os.path.join(ROOT, "results.json")
RESPONSES_DIR = os.path.join(ROOT, "responses")
IMAGES_DIR = os.path.join(ROOT, "images")
PROMPTS_DIR = os.path.join(ROOT, "prompts")
TREND_JSON = os.path.join(ROOT, "trend.json")

API_URL = "https://openrouter.ai/api/v1/chat/completions"

CATEGORIES = ["code", "text", "email", "image"]
EXCEL_CELL_LIMIT = 32000  # below Excel's hard 32,767 limit, with headroom

# In-flight counter for the concurrency sanity check (verification step 2).
_inflight_lock = threading.Lock()
_inflight = 0
_inflight_max = 0


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def log(msg):
    print(f"[{datetime.now(timezone.utc).strftime('%H:%M:%S')}] {msg}", flush=True)


def load_json(path, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def slugify(value):
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", value).strip("_") or "model"


def cat_setting(config, category, key, default=None):
    """Per-category override falling back to defaults block then a literal default."""
    cat = config.get("categories", {}).get(category, {})
    if key in cat:
        return cat[key]
    if key in config.get("defaults", {}):
        return config["defaults"][key]
    return default


# --------------------------------------------------------------------------- #
# OpenRouter call with retry / timeout
# --------------------------------------------------------------------------- #
def call_model(api_key, model, prompt, *, max_tokens, temperature, timeout,
               max_retries, backoff, want_image=False):
    """
    Returns a dict with metrics and either the text/image or an error.
    Retries on timeouts, 429 and 5xx with exponential backoff.
    """
    global _inflight, _inflight_max

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://github.com/model-bench",
        "X-Title": "model-bench",
    }
    body = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    if want_image:
        body["modalities"] = ["image", "text"]

    last_err = None
    for attempt in range(1, max_retries + 1):
        with _inflight_lock:
            _inflight += 1
            _inflight_max = max(_inflight_max, _inflight)
        start = time.monotonic()
        try:
            resp = requests.post(API_URL, headers=headers, json=body, timeout=timeout)
            latency_ms = int((time.monotonic() - start) * 1000)

            if resp.status_code == 200:
                data = resp.json()
                choice = (data.get("choices") or [{}])[0]
                message = choice.get("message", {}) or {}
                text = message.get("content") or ""
                usage = data.get("usage", {}) or {}

                images_b64 = []
                for img in message.get("images", []) or []:
                    url = (img.get("image_url") or {}).get("url", "")
                    if url.startswith("data:") and "," in url:
                        images_b64.append(url.split(",", 1)[1])

                return {
                    "status": "ok",
                    "http_status": 200,
                    "tokens_in": usage.get("prompt_tokens"),
                    "tokens_out": usage.get("completion_tokens"),
                    "total_tokens": usage.get("total_tokens"),
                    "latency_ms": latency_ms,
                    "text": text,
                    "images_b64": images_b64,
                    "error": "",
                    "attempts": attempt,
                }

            # Retryable HTTP statuses
            if resp.status_code == 429 or 500 <= resp.status_code < 600:
                last_err = f"HTTP {resp.status_code}: {resp.text[:200]}"
                if attempt < max_retries:
                    time.sleep(backoff * (2 ** (attempt - 1)))
                    continue
            # Non-retryable
            return {
                "status": "error",
                "http_status": resp.status_code,
                "tokens_in": None, "tokens_out": None, "total_tokens": None,
                "latency_ms": latency_ms,
                "text": "", "images_b64": [],
                "error": f"HTTP {resp.status_code}: {resp.text[:300]}",
                "attempts": attempt,
            }
        except (requests.Timeout, requests.ConnectionError) as e:
            last_err = f"{type(e).__name__}: {e}"
            if attempt < max_retries:
                time.sleep(backoff * (2 ** (attempt - 1)))
                continue
        except Exception as e:  # noqa: BLE001 - record anything unexpected, never crash the run
            last_err = f"{type(e).__name__}: {e}"
            break
        finally:
            with _inflight_lock:
                _inflight -= 1

    return {
        "status": "error",
        "http_status": None,
        "tokens_in": None, "tokens_out": None, "total_tokens": None,
        "latency_ms": None,
        "text": "", "images_b64": [],
        "error": last_err or "unknown error",
        "attempts": max_retries,
    }


# --------------------------------------------------------------------------- #
# Prompt sourcing
# --------------------------------------------------------------------------- #
def rotate(items, cycle):
    return items[cycle % len(items)] if items else None


def generate_prompts_batch(api_key, config, categories, generators, history, cycle):
    """
    Generate ALL category prompts in ONE API call (saves N-1 requests vs per-category).
    Returns {category: prompt_text}. Raises on failure.
    """
    parts = []
    for cat in categories:
        meta = generators.get(cat, "")
        recent = history.get(cat, [])[-config.get("prompt_history_keep", 20):]
        avoid = ""
        if recent:
            joined = "\n".join(f"    - {p[:200]}" for p in recent)
            avoid = f"\n  Recently used (do NOT repeat or closely resemble):\n{joined}"
        parts.append(f'  "{cat}": {meta}{avoid}')

    instruction = (
        f"Generate exactly one benchmark prompt for EACH of the following {len(categories)} categories. "
        f"Return ONLY a valid JSON object with the category name as the key and the complete prompt text as the value. "
        f"No markdown, no code fences, no text before or after the JSON. Variation seed: cycle {cycle}.\n\n"
        "Categories:\n" + "\n\n".join(parts)
    )
    result = call_model(
        api_key,
        config["generator_model"],
        instruction,
        max_tokens=3000,
        temperature=config.get("generator_temperature", 0.9),
        timeout=config["request_timeout_seconds"],
        max_retries=config["max_retries"],
        backoff=config["retry_backoff_seconds"],
    )
    if result["status"] != "ok" or not result["text"].strip():
        raise RuntimeError(result.get("error") or "empty generation")

    text = result["text"].strip()
    # Strip markdown code fences if the model added them
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
    # Find the JSON object
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        raise RuntimeError(f"no JSON object in generator response: {text[:300]}")
    parsed = json.loads(m.group())
    missing = [c for c in categories if c not in parsed or not str(parsed[c]).strip()]
    if missing:
        raise RuntimeError(f"missing categories in batch response: {missing}")
    return {c: str(parsed[c]).strip() for c in categories}


def resolve_prompt_single(api_key, config, category, generators, manual, fallback, history, cycle):
    """
    Single-category fallback used when batch generation is disabled or per-category
    prompt_source is 'manual' or 'fallback'.
    Returns (raw_prompt, source_used).
    """
    source = cat_setting(config, category, "prompt_source", "generate")

    if source == "manual":
        entry = manual.get(category)
        if isinstance(entry, list):
            entry = rotate(entry, cycle)
        if not entry:
            raise RuntimeError(f"no manual prompt for '{category}' in manual.json")
        return entry, "manual"

    if source == "fallback":
        entry = rotate(fallback.get(category, []), cycle)
        if not entry:
            raise RuntimeError(f"no fallback prompt for '{category}' in fallback.json")
        return entry, "fallback"

    # generate (single-category path — only used when batch is off for this cat)
    recent = history.get(category, [])[-config.get("prompt_history_keep", 20):]
    avoid = ""
    if recent:
        joined = "\n".join(f"- {p[:300]}" for p in recent)
        avoid = f"\n\nDo NOT repeat or closely resemble any of these:\n{joined}"
    instruction = (
        f"{generators.get(category,'')}\n\n(Variation seed: cycle {cycle}.)"
        f"{avoid}\n\nOutput ONLY the prompt text, no preamble."
    )
    result = call_model(
        api_key, config["generator_model"], instruction,
        max_tokens=2000,
        temperature=config.get("generator_temperature", 0.9),
        timeout=config["request_timeout_seconds"],
        max_retries=config["max_retries"],
        backoff=config["retry_backoff_seconds"],
    )
    if result["status"] != "ok" or not result["text"].strip():
        raise RuntimeError(result.get("error") or "empty generation")
    return result["text"].strip(), "generate"


def wrap_for_models(category, raw_prompt):
    """Turn the category content into the actual prompt sent to every model."""
    if category == "email":
        return (
            "You are a professional working in grants administration. Reply to the "
            "following email. Your reply must be professional, address every point "
            "raised, and be at least 200 words long.\n\n"
            f"--- EMAIL ---\n{raw_prompt}\n--- END EMAIL ---"
        )
    if category == "image":
        return f"Generate an image that visually explains this feature:\n\n{raw_prompt}"
    # code / text: the generated prompt is already the instruction
    return raw_prompt


# --------------------------------------------------------------------------- #
# Persistence
# --------------------------------------------------------------------------- #
EXCEL_HEADERS = [
    "run_id", "timestamp_utc", "cycle", "category", "prompt_source", "prompt_title",
    "model_id", "status", "http_status", "tokens_in", "tokens_out", "total_tokens",
    "latency_ms", "response_chars", "response_words", "attempts",
    "response_file", "image_file", "error",
]


def append_to_excel(rows):
    if os.path.exists(RESULTS_XLSX):
        wb = load_workbook(RESULTS_XLSX)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for category in CATEGORIES:
        cat_rows = [r for r in rows if r["category"] == category]
        if not cat_rows:
            continue
        if category in wb.sheetnames:
            ws = wb[category]
        else:
            ws = wb.create_sheet(title=category)
            ws.append(EXCEL_HEADERS)
        for r in cat_rows:
            ws.append([r.get(h, "") for h in EXCEL_HEADERS])

    if not wb.sheetnames:  # nothing ran; keep a valid workbook
        wb.create_sheet(title="empty")
    wb.save(RESULTS_XLSX)


def update_results_json(rows, run_meta, keep_runs):
    data = load_json(RESULTS_JSON, default={"runs": []})
    if not isinstance(data, dict):
        data = {"runs": []}
    compact = []
    for r in rows:
        compact.append({k: r.get(k) for k in (
            "category", "prompt_source", "prompt_title", "model_id", "status",
            "tokens_in", "tokens_out", "latency_ms", "response_chars",
            "response_words", "response_file", "image_file", "error",
        )})
    data["runs"].insert(0, {**run_meta, "results": compact})
    data["runs"] = data["runs"][:keep_runs]
    save_json(RESULTS_JSON, data)


def update_trend_json(rows, run_id, timestamp_utc, cycle, keep_runs):
    """
    Compact trend file: one entry per run containing only the metrics needed
    for charting (no prompt text, no error strings). Stays small forever.

    Structure: { "runs": [ { "run_id", "ts", "cycle", "points": [
        { "cat", "model", "ok", "tokens_out", "latency_ms", "response_chars" }, ...
    ] }, ... ] }
    """
    data = load_json(TREND_JSON, default={"runs": []})
    if not isinstance(data, dict):
        data = {"runs": []}

    points = []
    for r in rows:
        points.append({
            "cat":    r["category"],
            "model":  r["model_id"],
            "ok":     1 if r["status"] == "ok" else 0,
            "tout":   r["tokens_out"] if r["tokens_out"] != "" else None,
            "lat":    r["latency_ms"] if r["latency_ms"] != "" else None,
            "chars":  r["response_chars"] if r["response_chars"] else None,
        })

    data["runs"].append({
        "run_id": run_id,
        "ts":     timestamp_utc,
        "cycle":  cycle,
        "points": points,
    })
    # Keep oldest-first order (append to end); dashboard reads left-to-right as time.
    data["runs"] = data["runs"][-keep_runs:]
    save_json(TREND_JSON, data)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        log("ERROR: OPENROUTER_API_KEY not set")
        sys.exit(1)

    config = load_json(CONFIG_PATH)
    state = load_json(STATE_PATH, default={"cycle": 0, "last_run_utc": None})
    history = load_json(HISTORY_PATH, default={c: [] for c in CATEGORIES})
    generators = load_json(os.path.join(PROMPTS_DIR, "generators.json"), default={})
    manual = load_json(os.path.join(PROMPTS_DIR, "manual.json"), default={})
    fallback = load_json(os.path.join(PROMPTS_DIR, "fallback.json"), default={})

    cycle = int(state.get("cycle", 0))
    now = datetime.now(timezone.utc)
    run_id = now.strftime("%Y%m%dT%H%M%SZ")
    log(f"=== Run {run_id} (cycle {cycle}) ===")

    enabled = [c for c in CATEGORIES if config["categories"].get(c, {}).get("enabled")]
    log(f"Enabled categories: {enabled}")

    # 1) Resolve one prompt per enabled category.
    # Separate "generate" categories (batch into 1 API call) from manual/fallback (no call).
    prompts = {}  # category -> {raw, wrapped, source, title}
    gen_cats = [c for c in enabled
                if cat_setting(config, c, "prompt_source", "generate") == "generate"]
    other_cats = [c for c in enabled if c not in gen_cats]

    # --- Batch generate all "generate" categories in one call ---
    if gen_cats:
        try:
            batch = generate_prompts_batch(api_key, config, gen_cats, generators, history, cycle)
            for category, raw in batch.items():
                max_in = cat_setting(config, category, "max_input_chars", 8000)
                raw = raw[:max_in]
                prompts[category] = {"raw": raw, "source": "generate"}
            log(f"  [batch] generated prompts for: {list(batch.keys())}")
        except Exception as e:  # noqa: BLE001
            if config.get("generator_fallback_on_error", True):
                log(f"  ! batch generation failed ({e}); falling back per-category")
                for category in gen_cats:
                    entry = rotate(fallback.get(category, []), cycle)
                    if entry:
                        prompts[category] = {"raw": entry, "source": "fallback"}
                    else:
                        log(f"  ! no fallback for {category}, skipping")
            else:
                log(f"  ! batch generation failed ({e}); skipping {gen_cats}")

    # --- Manual / fallback categories (no API call) ---
    for category in other_cats:
        try:
            raw, source = resolve_prompt_single(
                api_key, config, category, generators, manual, fallback, history, cycle)
            prompts[category] = {"raw": raw, "source": source}
        except Exception as e:  # noqa: BLE001
            log(f"  ! skipping {category}: {e}")

    # --- Finalise prompts (wrap + title + history update) ---
    for category in list(prompts.keys()):
        p = prompts[category]
        wrapped = wrap_for_models(category, p["raw"])
        title = re.sub(r"\s+", " ", p["raw"]).strip()[:80]
        prompts[category] = {**p, "wrapped": wrapped, "title": title}
        history.setdefault(category, []).append(p["raw"])
        history[category] = history[category][-config.get("prompt_history_keep", 20):]
        log(f"  [{category}] prompt via {p['source']}: {title}")

    # 2) Build the work list: (category, model) -> one request.
    tasks = []
    for category, p in prompts.items():
        models = config["categories"].get(category, {}).get("models_override") or config["models"]
        for model in models:
            tasks.append((category, model, p))
    log(f"Dispatching {len(tasks)} requests across max {config['max_in_flight']} in flight")

    # 3) Execute with bounded concurrency.
    rows = []

    def run_task(task):
        category, model, p = task
        want_image = category == "image"
        res = call_model(
            api_key, model, p["wrapped"],
            max_tokens=cat_setting(config, category, "max_output_tokens", 4096),
            temperature=cat_setting(config, category, "temperature", 0.7),
            timeout=config["request_timeout_seconds"],
            max_retries=config["max_retries"],
            backoff=config["retry_backoff_seconds"],
            want_image=want_image,
        )
        return category, model, p, res

    with ThreadPoolExecutor(max_workers=config["max_in_flight"]) as pool:
        futures = [pool.submit(run_task, t) for t in tasks]
        for fut in as_completed(futures):
            category, model, p, res = fut.result()
            rows.append(persist_one(run_id, cycle, now, category, model, p, res, config))
            log(f"  done {category}/{model}: {res['status']} "
                f"({res.get('latency_ms')}ms, out={res.get('tokens_out')})")

    # 4) Write Excel + JSON + trend.
    if rows:
        append_to_excel(rows)
        update_results_json(
            rows,
            {"run_id": run_id, "timestamp_utc": now.isoformat(), "cycle": cycle,
             "prompts": {c: prompts[c]["title"] for c in prompts}},
            config.get("dashboard_history_runs", 500),
        )
        update_trend_json(
            rows, run_id, now.isoformat(), cycle,
            config.get("trend_history_runs", 200),
        )

    # 5) Advance state + history.
    state["cycle"] = cycle + 1
    state["last_run_utc"] = now.isoformat()
    save_json(STATE_PATH, state)
    save_json(HISTORY_PATH, history)

    log(f"Wrote {len(rows)} rows. Peak in-flight: {_inflight_max}. Done.")


def persist_one(run_id, cycle, now, category, model, p, res, config):
    """Write the full response to disk and return the metrics row."""
    slug = slugify(model)
    ext = "java" if category == "code" else "md"
    rel_dir = os.path.join("responses", run_id, category)
    abs_dir = os.path.join(ROOT, rel_dir)
    os.makedirs(abs_dir, exist_ok=True)

    response_file = ""
    text = res.get("text") or ""
    if text:
        store_limit = cat_setting(config, category, "max_response_store_chars", 200000)
        rel_path = os.path.join(rel_dir, f"{slug}.{ext}")
        with open(os.path.join(ROOT, rel_path), "w", encoding="utf-8") as f:
            header = (f"<!-- model: {model} | category: {category} | run: {run_id} -->\n"
                      f"<!-- PROMPT -->\n{p['wrapped']}\n<!-- RESPONSE -->\n")
            f.write(header + text[:store_limit])
        response_file = rel_path

    image_file = ""
    for i, b64 in enumerate(res.get("images_b64", [])):
        img_rel_dir = os.path.join("images", run_id, category)
        os.makedirs(os.path.join(ROOT, img_rel_dir), exist_ok=True)
        img_rel = os.path.join(img_rel_dir, f"{slug}_{i}.png")
        try:
            with open(os.path.join(ROOT, img_rel), "wb") as f:
                f.write(base64.b64decode(b64))
            image_file = img_rel
        except Exception as e:  # noqa: BLE001
            res["error"] = (res.get("error") or "") + f" [image decode: {e}]"

    return {
        "run_id": run_id,
        "timestamp_utc": now.isoformat(),
        "cycle": cycle,
        "category": category,
        "prompt_source": p["source"],
        "prompt_title": p["title"],
        "model_id": model,
        "status": res["status"],
        "http_status": res.get("http_status") or "",
        "tokens_in": res.get("tokens_in") if res.get("tokens_in") is not None else "",
        "tokens_out": res.get("tokens_out") if res.get("tokens_out") is not None else "",
        "total_tokens": res.get("total_tokens") if res.get("total_tokens") is not None else "",
        "latency_ms": res.get("latency_ms") if res.get("latency_ms") is not None else "",
        "response_chars": len(text),
        "response_words": len(text.split()),
        "attempts": res.get("attempts", 1),
        "response_file": response_file,
        "image_file": image_file,
        "error": (res.get("error") or "")[:EXCEL_CELL_LIMIT],
    }


if __name__ == "__main__":
    main()
